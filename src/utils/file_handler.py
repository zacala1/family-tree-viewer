"""파일 Import/Export 핸들러."""

import json
import os
from typing import Optional, TYPE_CHECKING, Any, Literal
from datetime import datetime

from .logger import error, warning

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# TYPE_CHECKING을 사용하여 순환 import 방지
if TYPE_CHECKING:
    from ..models.family_tree import FamilyTree


class FileHandler:
    """파일 Import/Export를 처리하는 클래스."""

    # 지원 파일 형식
    SUPPORTED_FORMATS = {
        "json": "Family Tree JSON (*.json)",
        "xlsx": "Excel Workbook (*.xlsx)",
        "gedcom": "GEDCOM (*.ged)",
    }

    @staticmethod
    def get_save_filters() -> str:
        """저장 다이얼로그용 필터 문자열."""
        return ";;".join(
            [
                "Family Tree JSON (*.json)",
                "Excel Workbook (*.xlsx)",
            ]
        )

    @staticmethod
    def get_open_filters() -> str:
        """열기 다이얼로그용 필터 문자열."""
        return ";;".join(
            [
                "All Supported Files (*.json *.xlsx *.ged)",
                "Family Tree JSON (*.json)",
                "Excel Workbook (*.xlsx)",
                "GEDCOM (*.ged)",
            ]
        )

    # === JSON ===

    @staticmethod
    def save_json(tree: "FamilyTree", file_path: str) -> bool:
        """JSON 형식으로 저장 (원자적 쓰기)."""
        import tempfile
        import shutil

        try:
            data = tree.to_dict()
            data["_meta"] = {
                "version": "1.0",
                "created": datetime.now().isoformat(),
                "app": "FamilyTree",
            }

            # 디렉토리가 존재하는지 확인
            dir_path = os.path.dirname(file_path)
            if dir_path and not os.path.exists(dir_path):
                os.makedirs(dir_path, exist_ok=True)

            # 임시 파일에 먼저 저장 (원자적 쓰기 + 백업)
            temp_fd, temp_path = tempfile.mkstemp(suffix=".json", dir=dir_path or ".")
            backup_path = file_path + ".backup" if os.path.exists(file_path) else None

            try:
                with os.fdopen(temp_fd, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)

                # 기존 파일 백업
                if backup_path:
                    shutil.copy2(file_path, backup_path)

                # Windows: 기존 파일이 있으면 삭제 필요
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except PermissionError as e:
                        error(f"Cannot remove existing file (file may be open): {file_path}")
                        raise PermissionError(f"File is in use: {file_path}") from e

                # 원자적 교체
                shutil.move(temp_path, file_path)

                # 성공 시 백업 삭제
                if backup_path and os.path.exists(backup_path):
                    os.remove(backup_path)

                return True
            except Exception as e:
                # 실패 시 백업에서 복원
                if backup_path and os.path.exists(backup_path):
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                        shutil.move(backup_path, file_path)
                        error(f"Save failed, restored from backup: {e}")
                    except Exception as restore_error:
                        error(f"Failed to restore from backup: {restore_error}")

                # 임시 파일 정리
                try:
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)
                except OSError as cleanup_error:
                    error(f"Failed to clean up temp file {temp_path}: {cleanup_error}")
                raise

        except PermissionError as e:
            error(f"Permission denied: {file_path} - {e}")
            return False
        except OSError as e:
            error(f"OS error while saving JSON: {e}")
            return False
        except (TypeError, ValueError) as e:
            error(f"Data serialization error: {e}")
            return False
        except Exception as e:
            error(f"Unexpected error saving JSON: {e}")
            return False

    @staticmethod
    def load_json(file_path: str) -> Optional["FamilyTree"]:
        """JSON 파일 로드."""
        from ..models.family_tree import FamilyTree

        try:
            if not os.path.exists(file_path):
                error(f"File not found: {file_path}")
                return None

            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            return FamilyTree.from_dict(data)
        except PermissionError:
            error(f"Permission denied: {file_path}")
            return None
        except json.JSONDecodeError as e:
            error(f"Invalid JSON format: {e}")
            return None
        except Exception as e:
            error(f"JSON load error: {e}")
            return None

    # === Excel ===

    @staticmethod
    def save_excel(tree: "FamilyTree", file_path: str) -> bool:
        """Excel 형식으로 저장."""
        if not HAS_OPENPYXL:
            error("openpyxl library is not installed")
            return False

        try:
            # 디렉토리가 존재하는지 확인
            dir_path = os.path.dirname(file_path)
            if dir_path and not os.path.exists(dir_path):
                os.makedirs(dir_path, exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "가족 구성원"

            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            # 헤더
            headers = [
                "ID",
                "이름",
                "성별",
                "출생연도",
                "출생월",
                "출생일",
                "음력여부",
                "사망연도",
                "사망월",
                "사망일",
                "출생지",
                "현주소",
                "직업",
                "학력",
                "연락처",
                "이메일",
                "메모",
                "아버지ID",
                "어머니ID",
                "배우자ID",
                "세대",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # 데이터
            for row, person in enumerate(tree.get_all_persons(), 2):
                ws.cell(row=row, column=1, value=person.id)
                ws.cell(row=row, column=2, value=person.name)
                ws.cell(row=row, column=3, value="남" if person.gender == "M" else "여")
                ws.cell(row=row, column=4, value=person.birth_year)
                ws.cell(row=row, column=5, value=person.birth_month)
                ws.cell(row=row, column=6, value=person.birth_day)
                ws.cell(row=row, column=7, value="예" if person.is_lunar_birth else "아니오")
                ws.cell(row=row, column=8, value=person.death_year)
                ws.cell(row=row, column=9, value=person.death_month)
                ws.cell(row=row, column=10, value=person.death_day)
                ws.cell(row=row, column=11, value=person.birth_place)
                ws.cell(row=row, column=12, value=person.current_address)
                ws.cell(row=row, column=13, value=person.occupation)
                ws.cell(row=row, column=14, value=person.education)
                ws.cell(row=row, column=15, value=person.phone)
                ws.cell(row=row, column=16, value=person.email)
                ws.cell(row=row, column=17, value=person.notes)
                ws.cell(row=row, column=18, value=person.father_id)
                ws.cell(row=row, column=19, value=person.mother_id)
                ws.cell(row=row, column=20, value=",".join(person.spouse_ids))
                ws.cell(row=row, column=21, value=person.generation)

            # 열 너비 자동 조정
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except (TypeError, ValueError):
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(file_path)
            return True

        except PermissionError:
            error(f"Permission denied: {file_path}")
            return False
        except OSError as e:
            error(f"OS error while saving Excel: {e}")
            return False
        except Exception as e:
            error(f"Excel save error: {e}")
            return False

    @staticmethod
    def _safe_int(value, default=None) -> Optional[int]:
        """안전하게 정수 변환."""
        if value is None or value == "":
            return default
        try:
            return int(float(value))  # Excel에서 숫자가 float으로 올 수 있음
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _safe_str(value, default="") -> str:
        """안전하게 문자열 변환."""
        if value is None:
            return default
        return str(value).strip()

    @staticmethod
    def load_excel(file_path: str) -> Optional["FamilyTree"]:
        """Excel 파일 로드."""
        from ..models.family_tree import FamilyTree
        from ..models.person import Person

        if not HAS_OPENPYXL:
            error("openpyxl library is not installed")
            return None

        try:
            if not os.path.exists(file_path):
                error(f"File not found: {file_path}")
                return None

            wb = load_workbook(file_path)
            ws = wb.active

            tree = FamilyTree()

            # 예상 컬럼 수
            EXPECTED_COLUMNS = 21

            # 헤더 건너뛰고 데이터 읽기
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 컬럼 수 검증
                    if not row or len(row) < EXPECTED_COLUMNS:
                        warning(
                            f"Excel row {row_num}: Insufficient columns (expected {EXPECTED_COLUMNS}, got {len(row) if row else 0}) - skipping"
                        )
                        continue

                    if not row[0]:  # ID가 없으면 건너뛰기
                        continue

                    gender_str = FileHandler._safe_str(row[2], "남")
                    gender: Literal["M", "F"] = "M" if gender_str == "남" else "F"

                    lunar_str = FileHandler._safe_str(row[6], "아니오")
                    is_lunar = lunar_str == "예"

                    spouse_ids = []
                    if row[19]:
                        spouse_ids = [s.strip() for s in str(row[19]).split(",") if s.strip()]

                    person = Person(
                        id=FileHandler._safe_str(row[0]),
                        name=FileHandler._safe_str(row[1]),
                        gender=gender,
                        birth_year=FileHandler._safe_int(row[3]),
                        birth_month=FileHandler._safe_int(row[4]),
                        birth_day=FileHandler._safe_int(row[5]),
                        is_lunar_birth=is_lunar,
                        death_year=FileHandler._safe_int(row[7]),
                        death_month=FileHandler._safe_int(row[8]),
                        death_day=FileHandler._safe_int(row[9]),
                        birth_place=FileHandler._safe_str(row[10]),
                        current_address=FileHandler._safe_str(row[11]),
                        occupation=FileHandler._safe_str(row[12]),
                        education=FileHandler._safe_str(row[13]),
                        phone=FileHandler._safe_str(row[14]),
                        email=FileHandler._safe_str(row[15]),
                        notes=FileHandler._safe_str(row[16]),
                        father_id=FileHandler._safe_str(row[17]) or None,
                        mother_id=FileHandler._safe_str(row[18]) or None,
                        spouse_ids=spouse_ids,
                        generation=FileHandler._safe_int(row[20], 0),
                    )

                    tree.add_person(person)

                except Exception as row_error:
                    warning(f"Excel row {row_num} load error: {row_error} - skipping")
                    continue

            # 부모 참조에서 children_ids 복구
            FileHandler._rebuild_children_ids(tree)

            tree.mark_saved()
            return tree

        except PermissionError:
            error(f"Permission denied: {file_path}")
            return None
        except Exception as e:
            error(f"Excel load error: {e}")
            return None

    @staticmethod
    def _rebuild_children_ids(tree: "FamilyTree") -> None:
        """부모 참조로부터 children_ids를 복구."""
        for person in tree.get_all_persons():
            if person.father_id:
                father = tree.get_person(person.father_id)
                if father and person.id not in father.children_ids:
                    father.children_ids.append(person.id)
            if person.mother_id:
                mother = tree.get_person(person.mother_id)
                if mother and person.id not in mother.children_ids:
                    mother.children_ids.append(person.id)

    # === GEDCOM ===

    @staticmethod
    def _parse_gedcom_line(line: str) -> Optional[tuple[int, str, str]]:
        """GEDCOM 라인 파싱.

        Returns:
            (level, tag, value) 또는 None
        """
        parts = line.split(" ", 2)
        if not parts:
            return None

        try:
            level = int(parts[0])
        except (ValueError, IndexError):
            warning(f"GEDCOM: Invalid line format (missing level): {line}")
            return None

        tag = parts[1] if len(parts) > 1 else ""
        value = parts[2] if len(parts) > 2 else ""
        return (level, tag, value)

    @staticmethod
    def _process_indi_record(
        tag: str, value: str, current_data: dict[str, Any]
    ) -> None:
        """INDI 레코드 처리."""
        if tag == "NAME":
            name = value.replace("/", "").strip()
            current_data["name"] = name
        elif tag == "SEX":
            current_data["gender"] = value
        elif tag == "BIRT":
            current_data["_in_birth"] = True
        elif tag == "DEAT":
            current_data["_in_death"] = True
        elif tag == "DATE" and current_data.get("_in_birth"):
            current_data["birth_date"] = value
            current_data["_in_birth"] = False
        elif tag == "DATE" and current_data.get("_in_death"):
            current_data["death_date"] = value
            current_data["_in_death"] = False

    @staticmethod
    def _process_fam_record(tag: str, value: str, current_data: dict[str, Any]) -> None:
        """FAM 레코드 처리."""
        if tag == "HUSB":
            current_data["husb"] = value
        elif tag == "WIFE":
            current_data["wife"] = value
        elif tag == "CHIL":
            if "children" not in current_data:
                current_data["children"] = []
            current_data["children"].append(value)

    @staticmethod
    def _create_persons_from_gedcom(
        persons: dict[str, Any], tree: "FamilyTree"
    ) -> dict[str, str]:
        """GEDCOM 데이터에서 Person 객체 생성 및 ID 맵핑."""
        from ..models.person import Person

        id_map = {}
        for ged_id, data in persons.items():
            gender_val = data.get("gender", "M")
            gender: Literal["M", "F"] = "M" if gender_val != "F" else "F"
            person = Person(name=data.get("name", ""), gender=gender)

            if "birth_date" in data:
                year = FileHandler._parse_gedcom_year(data["birth_date"])
                if year:
                    person.birth_year = year

            if "death_date" in data:
                year = FileHandler._parse_gedcom_year(data["death_date"])
                if year:
                    person.death_year = year

            tree.add_person(person)
            id_map[ged_id] = person.id

        return id_map

    @staticmethod
    def _create_relationships_from_gedcom(
        families: dict[str, Any], id_map: dict[str, str], tree: "FamilyTree"
    ) -> None:
        """GEDCOM 가족 데이터에서 관계 생성."""
        for fam_data in families.values():
            husb_id = id_map.get(fam_data.get("husb"))
            wife_id = id_map.get(fam_data.get("wife"))
            children = [
                id_map.get(c) for c in fam_data.get("children", []) if id_map.get(c)
            ]

            if husb_id and wife_id:
                tree.set_spouse(husb_id, wife_id)

            for child_id in children:
                if husb_id:
                    tree.set_parent_child(husb_id, child_id)
                if wife_id:
                    tree.set_parent_child(wife_id, child_id)

    @staticmethod
    def load_gedcom(file_path: str) -> Optional["FamilyTree"]:
        """GEDCOM 파일 로드 (기본 파싱, DOS 방지)."""
        from ..models.family_tree import FamilyTree

        MAX_FILE_SIZE = 100 * 1024 * 1024
        MAX_LINES = 1000000

        try:
            if not os.path.exists(file_path):
                error(f"File not found: {file_path}")
                return None

            file_size = os.path.getsize(file_path)
            if file_size > MAX_FILE_SIZE:
                error(f"GEDCOM file too large: {file_size} bytes (max {MAX_FILE_SIZE})")
                return None

            tree = FamilyTree()
            persons: dict[str, Any] = {}
            families: dict[str, Any] = {}

            current_record = None
            current_id = None
            current_data: dict[str, Any] = {}

            # 스트리밍 방식으로 라인 처리 (메모리 효율성)
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                for line_num, line in enumerate(f, start=1):
                    if line_num > MAX_LINES:
                        error(f"GEDCOM file has too many lines (max {MAX_LINES})")
                        return None

                    line = line.strip()
                    if not line:
                        continue

                    parsed = FileHandler._parse_gedcom_line(line)
                    if parsed is None:
                        continue

                    level, tag, value = parsed

                    # 새 레코드 시작
                    if level == 0:
                        if current_record == "INDI" and current_id:
                            persons[current_id] = current_data.copy()
                        elif current_record == "FAM" and current_id:
                            families[current_id] = current_data.copy()

                        current_data = {}

                        if tag.startswith("@") and value in ("INDI", "FAM"):
                            current_id = tag
                            current_record = value
                        else:
                            current_record = None
                            current_id = None

                    elif current_record == "INDI":
                        FileHandler._process_indi_record(tag, value, current_data)

                    elif current_record == "FAM":
                        FileHandler._process_fam_record(tag, value, current_data)

            # 마지막 레코드 저장
            if current_record == "INDI" and current_id:
                persons[current_id] = current_data
            elif current_record == "FAM" and current_id:
                families[current_id] = current_data

            # Person 객체 생성 및 관계 설정
            id_map = FileHandler._create_persons_from_gedcom(persons, tree)
            FileHandler._create_relationships_from_gedcom(families, id_map, tree)

            tree.mark_saved()
            return tree

        except PermissionError:
            error(f"Permission denied: {file_path}")
            return None
        except UnicodeDecodeError as e:
            error(f"Encoding error in GEDCOM file: {e}")
            return None
        except Exception as e:
            error(f"GEDCOM load error: {e}")
            return None

    @staticmethod
    def _parse_gedcom_year(date_str: str) -> Optional[int]:
        """GEDCOM 날짜에서 연도 추출."""
        import re

        match = re.search(r"\b(\d{4})\b", date_str)
        if match:
            return int(match.group(1))
        return None

    # === 자동 감지 로드 ===

    @staticmethod
    def load_file(file_path: str) -> Optional["FamilyTree"]:
        """파일 확장자에 따라 자동으로 적절한 로더 사용."""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".json":
            return FileHandler.load_json(file_path)
        elif ext == ".xlsx":
            return FileHandler.load_excel(file_path)
        elif ext == ".ged":
            return FileHandler.load_gedcom(file_path)
        else:
            error(f"Unsupported file format: {ext}")
            return None

    @staticmethod
    def save_file(tree: "FamilyTree", file_path: str) -> bool:
        """파일 확장자에 따라 자동으로 적절한 저장 방식 사용."""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".json":
            return FileHandler.save_json(tree, file_path)
        elif ext == ".xlsx":
            return FileHandler.save_excel(tree, file_path)
        else:
            # 기본은 JSON
            if not file_path.endswith(".json"):
                file_path += ".json"
            return FileHandler.save_json(tree, file_path)
